import string
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
import openpyxl as xl  # type: ignore
import form
import sendemail as email

excel_filepath = "data/db.xlsx"


def leer_excel(excel_filepath: str) -> Worksheet:
    excel = xl.load_workbook(filename=excel_filepath).active
    return excel


def verificar_excel(excel: Worksheet) -> None:
    if excel["J1"].value == "Estado":
        pass
    else:
        raise Exception(
            'El archivo excel no es el indicado. No se encontró "Estado" en la celda J1'
        )


def leer_fila(excel: Worksheet, fila: int) -> list:
    fila_leida = []
    columnas = list(string.ascii_uppercase[0:10])
    for columna in columnas:
        fila_leida.append(excel[f"{columna}{fila}"].value)
    return fila_leida


def contar_filas(excel: Worksheet) -> int:
    max_row = excel.max_row
    # openpyxl busca max_rows que están vacias.
    # para eso las borro con el siguiente ciclo while
    while excel[f"A{max_row}"].value is None:
        excel.delete_rows(max_row)
        max_row = excel.max_row
    # y devuelvo la cantidad de rows que tienen texto
    return max_row


def main() -> None:
    excel = leer_excel(excel_filepath)
    verificar_excel(excel)
    total_filas = contar_filas(excel)
    for fila in range(2, total_filas):
        fila_leida = leer_fila(excel, fila)
        proceso = fila_leida[0]
        observacion = fila_leida[1]
        tipo_de_riesgo = fila_leida[2]
        severidad = fila_leida[3]
        fecha_de_compromiso = fila_leida[5]
        responsable = fila_leida[6]
        email_para = fila_leida[8]
        estado = fila_leida[9]

        if estado == "Regularizado":
            form.launch_chrome()
            form.completar_form(
                proceso=proceso,
                tipo_de_riesgo=tipo_de_riesgo,
                severidad=severidad,
                responsable=responsable,
                fecha_de_compromiso=fecha_de_compromiso,
                observacion=observacion,
            )
            form.driver.quit()
        elif estado == "Atrasado":
            email.configurar_servidor()
            email.enviar_email(
                email_para=email_para,
                proceso=proceso,
                estado=estado,
                fecha_de_compromiso=fecha_de_compromiso,
                observacion=observacion,
            )
            email.cerrar_servidor()
        elif estado == "Pendiente":
            # no se hace nada, continuar
            continue
        else:
            raise Exception(f"No se reconoce el estado {estado} en la fila {fila}")


if __name__ == "__main__":
    main()

"""
Limpiar variables
leer_excel()
    if abrir_archivo = excel:
        leer_excel
    else: 
        "no se reconoce archivo excel"

verificar_excel()
    if celda J1 == 'estado'
        pass
    else:
        'el archivo excel no es el indicado'
    
leer fila:
    contar filas de archivo

    for fila in filas_archivo
        leer una fila
        crear dict
        devolver una lista que es la fila de excel
crear dict
    asignar proceso a el item 0
    asignar observacion a el item 1
    asignar tipo_de_riesgo a el item 2
    asignar severidad a el item 3
    asignar fecha_de_compromiso a el item 5
    # acá divido la fecha en %d %m %Y
    asignar responsable a el item 6
    asignar email_para a el item 8
    asignar estado a el item 9

    if estado == "Regularizado"
        form.py con parametros de la fila
    elif estado == "Atrasado"
        email.py con parametros de la fila
    elif estado == "Pendiente"
        continue
    else:
        print('no se reconoce el estado de la fila {fila}')


"""
